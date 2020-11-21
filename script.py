import os
import openpyxl
import base64
from datetime import datetime
import xmlrpc.client
import psycopg2

workbook = openpyxl.load_workbook(filename='excel_file.xlsx')
worksheet = workbook.active
MAX_ROW = 100 #worksheet.max_row

url = 'http://localhost:8000'
db = 'inventory_test_db'
username = 'admin'
password = 'admin'

common = xmlrpc.client.ServerProxy('{}/xmlrpc/2/common'.format(url))
uid = common.authenticate(db, username, password, {})
models = xmlrpc.client.ServerProxy('{}/xmlrpc/2/object'.format(url))

def db_connection():
    
    USER = 'ubuntu'
    PASSWORD = '1234'
    HOST = '127.0.0.1'
    PORT = ''
    DATABASE = 'inventory_test_db'
    connection = None

    try:
        connection = psycopg2.connect(user=USER, password=PASSWORD, host=HOST, port=PORT, database=DATABASE)
    except (Exception, psycopg2.Error) as error:
        print('Error while connection to postgres', error)

    return connection

connection = db_connection()
cursor = connection.cursor()

def get_commodity_group_id(commodity_value):
    if commodity_value:
        cursor.execute("""SELECT id FROM product_commodity_group WHERE name = %s """ ,(commodity_value,))
        commodity_group_id = cursor.fetchone()
        return commodity_group_id
    else:
        return None

def encode_image(image_value):
    try:
        if image_value:
            img = open('photos/{img}'.format(img=image_value), 'rb')
            im_b64 = base64.b64encode(img.read()).decode('utf-8')
            return im_b64
        else:
            return ''
    except:
        return ''

start_reading_time = datetime.now()

print('================START=========================')

for row in range(2, MAX_ROW):
    item = worksheet.cell(row=row, column=1).value
    description = worksheet.cell(row=row, column=2).value
    part_number = worksheet.cell(row=row, column=3).value
    commodity_group_id = get_commodity_group_id(worksheet.cell(row=row, column=4).value)
    encoded_img = encode_image(worksheet.cell(row=row, column=5).value)    
    id = models.execute_kw(db, uid, password, 'product.template', 'create', [{
            'default_code': item,
            'name': description,
            'commodity_group': commodity_group_id[0] if commodity_group_id else 0,
            'image_1920': encoded_img,
        }])
            
print('===================END=======================')
print(datetime.now() - start_reading_time)




